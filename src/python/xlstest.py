#!/usr/bin/env python3

import sys,os,getopt,glob,math
import locale
import numpy as np
import xlrd
import openpyxl
import pandas as pd

def isFloat(string):
  try:
    float(string)
    return True
  except ValueError:
    return False

def usage():
  str = "Usage: python xlstest.py " \
        " --indir=? -outdir=?\n" \
        " --file=?\n" \
        "   indir: a directory containing input files to be processed\n" \
        "   outdir: a directory containing output files\n" \
        "   file: a single file to process\n"
  print(str)

def xlstest_pandas(infile,sheet):

# using pandas
  df = pd.read_excel(infile)
  print(df)
 
  for index, row in df.iterrows():
    print(index, row)

  # incomplete...I didn't bother with this once I got xlrd working
  # then they broke xlrd because they pulled support for reading xlsx files

  data = None

  return data

def xlstest_xlrd(infile,sheet):

# using xlrd (which dropped support for xlsx, only xls works now)
  workbook = xlrd.open_workbook(infile)

  infile = infile.replace('\\','/')
	
  # get infile base 
# base = infile[:infile.rfind('.')]
  base = os.path.basename(infile)
  print("Processing: ", infile, "[", base, "]")

  # strip out extension
  filename, ext = os.path.splitext(base)
  print('Processing: ' + filename)

  outfile = None

  # get worksheet by index
  print('Processing sheet: ' + str(sheet))

# using xlrd
  worksheet = workbook.sheet_by_index(sheet)

  # top row is header
  header = []

# using xlrd
  for col in range(worksheet.ncols):
    header.append(worksheet.cell_value(0,col))

  print(header)

  # using xlrd
  # transform workbook to a list of dicts
  data = []
  for row in range(1,worksheet.nrows):
    # set up element as dict so that we use column heading as key
    elem = {}
    for col in range(worksheet.ncols):
      elem[header[col]] = worksheet.cell_value(row,col)
#   print(elem)
    data.append(elem)

  return data

def xlstest_openpyxl(infile,sheet):

# using pyopenpyxl
  workbook = openpyxl.load_workbook(infile)
  print("Sheetname: ", workbook.sheetnames[sheet])

  infile = infile.replace('\\','/')
	
  # get infile base 
# base = infile[:infile.rfind('.')]
  base = os.path.basename(infile)
  print("Processing: ", infile, "[", base, "]")

  # strip out extension
  filename, ext = os.path.splitext(base)
  print('Processing: ' + filename)

  outfile = None

  # get worksheet by index
  print('Processing sheet: ' + str(sheet))

# using openpyxl
  worksheet = workbook[workbook.sheetnames[sheet]]
  print("worksheet dimensions: ", worksheet.dimensions)
  print("worksheet min_row: ", worksheet.min_row)
  print("worksheet max_row: ", worksheet.max_row)
  print("worksheet min_column: ", worksheet.min_column)
  print("worksheet max_column: ", worksheet.max_column)

  # top row is header
  header = []

# using openpyxl
  # warning: dimensions are 1..n not 0..n-1, so we need to go one extra col
  for col in range(worksheet.min_column,worksheet.max_column+1):
    header.append(worksheet.cell(row=worksheet.min_row,column=col).value)

  print(header)

  # using openpyxl
  # transform workbook to a list of dicts
  data = []
  # skip first row; also worksheet dimensions are 1..n, not 0..n-1
  for i in range(worksheet.min_row+1,worksheet.max_row+1):
    # set up element as dict so that we use column heading as key
    elem = {}
    for j in range(worksheet.min_column,worksheet.max_column):
      elem[header[j-1]] = worksheet.cell(row=i,column=j).value
#   print(elem)
    data.append(elem)

  return data

def dump(data,sheet):

# print(data)
  print("Number of entries: ",len(data))

  # recd is dict of dicts where we use column 'ido' as key
  # we use attr as the dict that we store for each recd
  recd = {}

  # process each line
  for i, entry in enumerate(data):

    # this gives strings
    # use eval() to convert tuple string to tuple
    ido = entry['ID okulo']
    idd = entry['ID docelowe ?']
    skl = entry['szkoła 0-m, 1-nm, 2-sportowa']

    # dict to store above key, val pairs
    attr = {}

    if ido is '':
      print("skipping %d" % (i))
    else:
#     print("%d,ido,idd: %s, %s" % (sheet + 2,ido,idd))
      attr['pomiar'] = sheet + 2
      attr['subjid'] = idd
      attr['szkola'] = skl
      # here's how we store the dict into a dict
      recd[ido] = attr

  print("Number of entries: ",len(recd))
# print(recd)
  # here's how we want to transofrm the file name as per the Excel sheet
  for k, v in recd.items():
    if k is not None:
#     print(k,v)
      print("%s -> %d_%s_%d" % (k,v['pomiar'],v['subjid'],v['szkola']))

def main(argv):
# if not len(argv):
#   usage()

  try:
    opts, args = getopt.getopt(argv, '', \
                 ['indir=','outdir=','file=',\
                  'hertz=','sfdegree=','sfcutoff='])
  except getopt.GetoptError as err:
    print(err)
    usage()
    exit()

  file = None
  files = []

  indir = '../../data/'
  file = indir + 'ID czytanie I i II pomiar.xlsx'
# file = indir + 'ID czytanie I i II pomiar.xls'
  outdir = './'

  for opt,arg in opts:
    opt = opt.lower()
    if(opt != '--file'):
      arg = arg.lower()

    if opt == '--indir':
      indir = arg
    elif opt == '--outdir':
      outdir = arg
    else:
      sys.argv[1:]

  print("file: ", file)

  # if user specified --file="..." then we use that as the only one to process
  if(file != None and os.path.isfile(file)):
    files = [file]
  elif os.path.isdir(indir):
    # get .csv input files to process
    for top, dirs, listing in os.walk(indir):
      for file in listing:
        path = os.path.join(top,file)
        dname = path.split('/')[2]
        base = os.path.basename(path)
        if base.endswith('.xlsx'):
          basename = base[:base.rfind('.xlsx')]
#         print(path, top, dname, base, "[",basename,"]","[",file,"]")
          files.extend([path])

  print("files: ", files)

  for file in files:
    print("Opening: ",file)
    for i in range(2):
#     data = xlstest_xlrd(file,i)
      data = xlstest_openpyxl(file,i)
      dump(data,i)

#######################
locale.setlocale( locale.LC_ALL, 'en_US.UTF-8' )

if __name__ == "__main__":
  main(sys.argv[1:])
